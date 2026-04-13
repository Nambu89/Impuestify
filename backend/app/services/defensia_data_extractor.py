"""Extractor de datos estructurados de documentos tributarios.

Cada función `extract_*` devuelve un dict tipado con los campos relevantes
del documento. Las funciones usan Gemini Vision internamente y exponen
helpers `_gemini_extract_*` aislados para facilitar mocks en tests.

Cálculos derivados (ej: diff_gastos_adquisicion_no_admitidos) se hacen en
Python puro sobre los campos devueltos por Gemini — cero LLM en esa parte,
para trazabilidad.
"""
from __future__ import annotations
import json
import logging
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def _parse_gemini_json(raw: str) -> dict[str, Any]:
    """Limpia y parsea la respuesta JSON de Gemini.

    Gemini a veces envuelve el JSON en ```json ... ``` o añade texto extra.
    Esta función aisla la lógica de limpieza para que sea consistente entre
    extractores y fácil de testear.
    """
    cleaned = raw.strip().strip("`")
    if cleaned.startswith("json"):
        cleaned = cleaned[4:].lstrip()
    return json.loads(cleaned)


_PROMPT_LIQUIDACION = """Eres un extractor de datos de liquidaciones provisionales
de IRPF emitidas por la AEAT española. Dado el texto del documento, devuelve un
JSON con los siguientes campos EXACTOS (null si no aparecen):

{
  "referencia": string,
  "fecha_acto": "YYYY-MM-DD",
  "cuota": number,
  "intereses_demora": number,
  "total_a_ingresar": number,
  "ejercicio": integer,
  "ccaa": string (nombre canónico con tildes),
  "tipo_tributo": "IRPF",
  "plazo_recurso_dias": integer,
  "ganancia_patrimonial": number | null,
  "gastos_adquisicion_declarados": number | null,
  "gastos_adquisicion_admitidos": number | null,
  "gastos_transmision_admitidos": number | null,
  "motivacion_articulos_citados": [string]
}

NO inventes valores. Si un campo no aparece literalmente, devuelve null.
Responde SOLO con el JSON, sin texto adicional.

DOCUMENTO:
"""


def _gemini_extract_liquidacion(pdf_bytes: bytes, nombre: str) -> dict[str, Any]:
    """Llama a Gemini Vision sobre el PDF. Aislado para mock en tests."""
    from google import genai
    from app.config import settings

    client = genai.Client(api_key=settings.GOOGLE_GEMINI_API_KEY)
    response = client.models.generate_content(
        model="gemini-3-flash-preview",
        contents=[
            _PROMPT_LIQUIDACION,
            {"inline_data": {"mime_type": "application/pdf", "data": pdf_bytes}},
        ],
    )
    return _parse_gemini_json(response.text)


def extract_liquidacion_provisional(pdf_bytes: bytes, nombre: str) -> dict[str, Any]:
    """Extrae datos de una liquidación provisional IRPF.

    Añade campos derivados calculados en Python puro:
    - diff_gastos_adquisicion_no_admitidos: diferencia entre gastos declarados
      y admitidos por AEAT (> 0 indica un posible argumento por falta de
      motivación específica).
    """
    try:
        datos = _gemini_extract_liquidacion(pdf_bytes, nombre)
    except Exception as exc:
        logger.error("Extracción liquidación falló para %s: %s", nombre, exc)
        return {"error": str(exc), "nombre": nombre}

    declarados = datos.get("gastos_adquisicion_declarados")
    admitidos = datos.get("gastos_adquisicion_admitidos")
    if declarados is not None and admitidos is not None:
        datos["diff_gastos_adquisicion_no_admitidos"] = round(
            declarados - admitidos, 2
        )

    return datos


_PROMPT_SANCION = """Eres un extractor de datos de acuerdos de imposición de
sanción tributaria de la AEAT española. Devuelve JSON EXACTO (null si no aparece):

{
  "referencia": string,
  "fecha_acto": "YYYY-MM-DD",
  "importe_sancion": number,
  "base_sancion_191": number | null,
  "porcentaje_191": number | null,
  "calificacion_191": "leve"|"grave"|"muy grave"|null,
  "base_sancion_194": number | null,
  "porcentaje_194": number | null,
  "calificacion_194": "leve"|"grave"|"muy grave"|null,
  "articulos_tipicos": [string],
  "reducciones_aplicadas": number,
  "motivacion_culpabilidad": string,
  "plazo_recurso_dias": integer
}

No inventes valores. Responde solo con el JSON.

DOCUMENTO:
"""


def _gemini_extract_sancion(pdf_bytes: bytes, nombre: str) -> dict[str, Any]:
    """Llama a Gemini para extraer datos de sanción. Aislado para mock."""
    from google import genai
    from app.config import settings

    client = genai.Client(api_key=settings.GOOGLE_GEMINI_API_KEY)
    response = client.models.generate_content(
        model="gemini-3-flash-preview",
        contents=[
            _PROMPT_SANCION,
            {"inline_data": {"mime_type": "application/pdf", "data": pdf_bytes}},
        ],
    )
    return _parse_gemini_json(response.text)


def extract_acuerdo_sancion(pdf_bytes: bytes, nombre: str) -> dict[str, Any]:
    """Extrae datos de un acuerdo de imposición de sanción tributaria.

    Añade el derivado `tiene_doble_tipicidad_191_194` que alimenta la regla
    R006 (non bis in idem parcial entre Art. 191 y Art. 194.1 LGT).
    """
    try:
        datos = _gemini_extract_sancion(pdf_bytes, nombre)
    except Exception as exc:
        logger.error("Extracción sanción falló para %s: %s", nombre, exc)
        return {"error": str(exc), "nombre": nombre}

    tiene_191 = datos.get("base_sancion_191") is not None
    tiene_194 = datos.get("base_sancion_194") is not None
    datos["tiene_doble_tipicidad_191_194"] = tiene_191 and tiene_194

    return datos


_PROMPT_PROPUESTA = """Extractor de propuestas de liquidación provisional IRPF/IVA/ISD/ITP/PLUSVALIA.
Devuelve JSON EXACTO (null si no aparecen):
{
  "referencia": string,
  "fecha_acto": "YYYY-MM-DD",
  "plazo_alegaciones_dias": integer,
  "cuota_propuesta": number,
  "ejercicio": integer,
  "tipo_tributo": "IRPF"|"IVA"|"ISD"|"ITP"|"PLUSVALIA",
  "ajustes_propuestos": [{"concepto": string, "ajuste": string}]
}
No inventes valores. Solo el JSON.

DOCUMENTO:
"""


def _gemini_extract_propuesta(pdf_bytes: bytes, nombre: str) -> dict[str, Any]:
    from google import genai
    from app.config import settings

    client = genai.Client(api_key=settings.GOOGLE_GEMINI_API_KEY)
    response = client.models.generate_content(
        model="gemini-3-flash-preview",
        contents=[
            _PROMPT_PROPUESTA,
            {"inline_data": {"mime_type": "application/pdf", "data": pdf_bytes}},
        ],
    )
    return _parse_gemini_json(response.text)


def extract_propuesta_liquidacion(pdf_bytes: bytes, nombre: str) -> dict[str, Any]:
    """Extrae datos de una propuesta de liquidación (fase previa al acuerdo)."""
    try:
        return _gemini_extract_propuesta(pdf_bytes, nombre)
    except Exception as exc:
        logger.error("Extracción propuesta falló para %s: %s", nombre, exc)
        return {"error": str(exc), "nombre": nombre}


_PROMPT_REQUERIMIENTO = """Extractor de requerimientos tributarios AEAT.
Devuelve JSON EXACTO (null si no aparecen):
{
  "referencia": string,
  "fecha_acto": "YYYY-MM-DD",
  "plazo_aportar_docs_dias": integer,
  "documentacion_solicitada": [string],
  "ejercicio": integer,
  "tipo_procedimiento": "verificacion_datos"|"comprobacion_limitada"|"otros",
  "alcance": string
}
No inventes valores. Solo JSON.

DOCUMENTO:
"""


def _gemini_extract_requerimiento(pdf_bytes: bytes, nombre: str) -> dict[str, Any]:
    from google import genai
    from app.config import settings

    client = genai.Client(api_key=settings.GOOGLE_GEMINI_API_KEY)
    response = client.models.generate_content(
        model="gemini-3-flash-preview",
        contents=[
            _PROMPT_REQUERIMIENTO,
            {"inline_data": {"mime_type": "application/pdf", "data": pdf_bytes}},
        ],
    )
    return _parse_gemini_json(response.text)


def extract_requerimiento(pdf_bytes: bytes, nombre: str) -> dict[str, Any]:
    """Extrae datos de un requerimiento AEAT (fase inicial del procedimiento)."""
    try:
        return _gemini_extract_requerimiento(pdf_bytes, nombre)
    except Exception as exc:
        logger.error("Extracción requerimiento falló para %s: %s", nombre, exc)
        return {"error": str(exc), "nombre": nombre}


_PROMPT_ESCRITO_USUARIO = """Extractor de escritos presentados por el contribuyente
(alegaciones, recurso de reposición, reclamación TEAR, ampliación de alegaciones).
Devuelve JSON EXACTO (null si no aparecen):
{
  "tipo_escrito": "alegaciones"|"reposicion"|"reclamacion_tear"|"ampliacion_alegaciones"|"otros",
  "referencia_acto_impugnado": string,
  "fecha_presentacion": "YYYY-MM-DD",
  "organo_destinatario": string,
  "pretension_principal": string,
  "argumentos_invocados": [string],
  "tributo": "IRPF"|"IVA"|"ISD"|"ITP"|"PLUSVALIA"|null,
  "ejercicio": integer | null
}
No inventes valores. Solo JSON.

DOCUMENTO:
"""


def _gemini_extract_escrito_usuario(pdf_bytes: bytes, nombre: str) -> dict[str, Any]:
    from google import genai
    from app.config import settings

    client = genai.Client(api_key=settings.GOOGLE_GEMINI_API_KEY)
    response = client.models.generate_content(
        model="gemini-3-flash-preview",
        contents=[
            _PROMPT_ESCRITO_USUARIO,
            {"inline_data": {"mime_type": "application/pdf", "data": pdf_bytes}},
        ],
    )
    return _parse_gemini_json(response.text)


def extract_escrito_usuario(pdf_bytes: bytes, nombre: str) -> dict[str, Any]:
    """Extrae datos de un escrito ya presentado por el usuario.

    Los escritos previos del usuario son clave para la detección de fase
    procesal: si existe un escrito de reclamación TEAR posterior a una
    liquidación, la fase ya no es 'plazo de recurso' sino 'TEAR interpuesta'.
    """
    try:
        return _gemini_extract_escrito_usuario(pdf_bytes, nombre)
    except Exception as exc:
        logger.error("Extracción escrito usuario falló para %s: %s", nombre, exc)
        return {"error": str(exc), "nombre": nombre}


def extract_libro_registro_xlsx(xlsx_bytes: bytes, nombre: str) -> dict[str, Any]:
    """Extrae datos de un libro registro de facturas en Excel.

    Detecta heurísticamente columnas 'Base', 'IVA' y 'Total' por su nombre y
    calcula agregados. No usa LLM — todo determinista.

    Esto es una de las ventajas de precisión frente a ChatGPT: los datos
    estructurados del libro registro entran al motor de reglas sin pasar por
    extracción estadística de un LLM.
    """
    try:
        wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)
    except Exception as exc:
        logger.error("No se pudo abrir %s: %s", nombre, exc)
        return {"error": str(exc), "nombre": nombre}

    hojas = []
    total_bases = 0.0
    total_iva = 0.0

    for ws in wb.worksheets:
        filas_raw = list(ws.iter_rows(values_only=True))
        if not filas_raw:
            continue
        columnas = [str(c) if c is not None else "" for c in filas_raw[0]]
        filas = []
        for fila in filas_raw[1:]:
            if all(v is None for v in fila):
                continue
            registro = {columnas[i]: fila[i] for i in range(len(columnas))}
            filas.append(registro)
            for key, val in registro.items():
                if isinstance(val, (int, float)):
                    kl = key.lower()
                    if "base" in kl:
                        total_bases += float(val)
                    elif "iva" in kl:
                        total_iva += float(val)
        hojas.append({
            "nombre": ws.title,
            "columnas": columnas,
            "num_filas": len(filas),
            "filas": filas,
        })

    return {
        "hojas": hojas,
        "total_importe_bases": round(total_bases, 2),
        "total_importe_iva": round(total_iva, 2),
    }
