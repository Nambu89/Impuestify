"""Export accounting books to CSV and Excel formats.

Supports Libro Diario, Libro Mayor, Libro Registro de Facturas,
and PyG (Cuenta de Perdidas y Ganancias).
"""

from __future__ import annotations

import csv
import io
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    _HAS_OPENPYXL = True
except ImportError:  # pragma: no cover
    _HAS_OPENPYXL = False


def _require_openpyxl() -> None:
    if not _HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel exports. "
            "Install it with: pip install openpyxl"
        )


# ---------------------------------------------------------------------------
# Header constants
# ---------------------------------------------------------------------------

_DIARIO_HEADERS = [
    "Fecha", "N Asiento", "Cuenta", "Nombre Cuenta",
    "Debe", "Haber", "Concepto",
]

_MAYOR_HEADERS = [
    "Cuenta", "Nombre", "Total Debe", "Total Haber", "Saldo",
]

_REGISTRO_HEADERS = [
    "Fecha", "N Factura", "Tipo", "NIF Emisor", "Emisor",
    "NIF Receptor", "Receptor", "Base Imponible", "Tipo IVA %",
    "Cuota IVA", "Retencion IRPF", "Total", "Cuenta PGC", "Descripcion",
]

_DIARIO_KEYS = [
    "fecha", "num_asiento", "cuenta", "nombre_cuenta",
    "debe", "haber", "concepto",
]

_MAYOR_KEYS = [
    "cuenta", "nombre", "total_debe", "total_haber", "saldo",
]

_REGISTRO_KEYS = [
    "fecha_factura", "numero_factura", "tipo", "emisor_nif", "emisor_nombre",
    "receptor_nif", "receptor_nombre", "base_imponible", "tipo_iva",
    "cuota_iva", "retencion_irpf", "total", "cuenta_pgc", "concepto",
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _dicts_to_csv(
    rows: list[dict[str, Any]],
    headers: list[str],
    keys: list[str],
) -> bytes:
    """Serialize a list of dicts to UTF-8 CSV bytes (with BOM for Excel)."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for row in rows:
        writer.writerow([row.get(k, "") for k in keys])
    return buf.getvalue().encode("utf-8-sig")


def _write_header_row(ws: Any, headers: list[str]) -> None:
    """Write bold header row to an openpyxl worksheet."""
    bold = Font(bold=True)
    fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold
        cell.fill = fill


def _dicts_to_excel(
    rows: list[dict[str, Any]],
    headers: list[str],
    keys: list[str],
    sheet_name: str,
) -> bytes:
    """Serialize a list of dicts to an Excel workbook (single sheet)."""
    _require_openpyxl()
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    _write_header_row(ws, headers)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, key in enumerate(keys, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

class ContabilidadExportService:
    """Static methods to export accounting data to CSV / Excel."""

    # -- CSV exports --------------------------------------------------------

    @staticmethod
    def libro_diario_to_csv(entries: list[dict]) -> bytes:
        """Export Libro Diario entries to CSV bytes (UTF-8 with BOM)."""
        return _dicts_to_csv(entries, _DIARIO_HEADERS, _DIARIO_KEYS)

    @staticmethod
    def libro_mayor_to_csv(mayor: list[dict]) -> bytes:
        """Export Libro Mayor summaries to CSV bytes."""
        return _dicts_to_csv(mayor, _MAYOR_HEADERS, _MAYOR_KEYS)

    @staticmethod
    def libro_registro_to_csv(facturas: list[dict]) -> bytes:
        """Export Libro Registro de Facturas to CSV bytes."""
        return _dicts_to_csv(facturas, _REGISTRO_HEADERS, _REGISTRO_KEYS)

    # -- Excel exports ------------------------------------------------------

    @staticmethod
    def libro_diario_to_excel(entries: list[dict]) -> bytes:
        """Export Libro Diario entries to an Excel workbook."""
        return _dicts_to_excel(entries, _DIARIO_HEADERS, _DIARIO_KEYS, "Libro Diario")

    @staticmethod
    def libro_mayor_to_excel(mayor: list[dict]) -> bytes:
        """Export Libro Mayor summaries to an Excel workbook."""
        return _dicts_to_excel(mayor, _MAYOR_HEADERS, _MAYOR_KEYS, "Libro Mayor")

    @staticmethod
    def pyg_to_excel(pyg: dict) -> bytes:
        """Export Cuenta de Perdidas y Ganancias to Excel.

        Expected *pyg* dict keys:
            - ingresos: list[dict] with {concepto, importe}
            - gastos: list[dict] with {concepto, importe}
            - total_ingresos: float
            - total_gastos: float
            - resultado: float
            - year: int | str
            - disclaimer: str (optional)
        """
        _require_openpyxl()

        wb = Workbook()
        ws = wb.active
        ws.title = "PyG"

        bold = Font(bold=True)
        bold_big = Font(bold=True, size=14)
        currency_fmt = '#,##0.00 €'
        fill_header = PatternFill(
            start_color="D9E1F2", end_color="D9E1F2", fill_type="solid",
        )

        row = 1

        # Title
        ws.cell(row=row, column=1, value=f"Cuenta de Perdidas y Ganancias {pyg.get('year', '')}").font = bold_big
        row += 2  # blank row

        # --- INGRESOS ---
        ws.cell(row=row, column=1, value="INGRESOS").font = bold
        ws.cell(row=row, column=1).fill = fill_header
        ws.cell(row=row, column=2, value="Importe").font = bold
        ws.cell(row=row, column=2).fill = fill_header
        row += 1

        for item in pyg.get("ingresos", []):
            ws.cell(row=row, column=1, value=item.get("concepto", ""))
            cell = ws.cell(row=row, column=2, value=item.get("importe", 0))
            cell.number_format = currency_fmt
            row += 1

        ws.cell(row=row, column=1, value="Total Ingresos").font = bold
        cell = ws.cell(row=row, column=2, value=pyg.get("total_ingresos", 0))
        cell.font = bold
        cell.number_format = currency_fmt
        row += 2  # blank row

        # --- GASTOS ---
        ws.cell(row=row, column=1, value="GASTOS").font = bold
        ws.cell(row=row, column=1).fill = fill_header
        ws.cell(row=row, column=2, value="Importe").font = bold
        ws.cell(row=row, column=2).fill = fill_header
        row += 1

        for item in pyg.get("gastos", []):
            ws.cell(row=row, column=1, value=item.get("concepto", ""))
            cell = ws.cell(row=row, column=2, value=item.get("importe", 0))
            cell.number_format = currency_fmt
            row += 1

        ws.cell(row=row, column=1, value="Total Gastos").font = bold
        cell = ws.cell(row=row, column=2, value=pyg.get("total_gastos", 0))
        cell.font = bold
        cell.number_format = currency_fmt
        row += 2  # blank row

        # --- RESULTADO ---
        ws.cell(row=row, column=1, value="RESULTADO DEL EJERCICIO").font = bold_big
        cell = ws.cell(row=row, column=2, value=pyg.get("resultado", 0))
        cell.font = bold_big
        cell.number_format = currency_fmt
        row += 2

        # --- Disclaimer ---
        disclaimer = pyg.get("disclaimer", "")
        if disclaimer:
            ws.cell(row=row, column=1, value=disclaimer).font = Font(italic=True, color="888888")

        # Auto-width columns
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 18

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
