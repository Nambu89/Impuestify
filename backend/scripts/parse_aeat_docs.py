#!/usr/bin/env python3
"""
parse_aeat_docs.py — AEAT document parser for TaxIA reference files.

Subcommands:
  xsd        Parse Renta2024.xsd → data/reference/renta_2024_schema.json
  xls        Parse Modelo 130/131 XLS/XLSX → data/reference/modelo_13X_fields.json
  verifactu  Parse VeriFactu XSD/WSDL → docs/aeat/verifactu/*_reference.txt
  all        Run all three parsers

Usage:
  python backend/scripts/parse_aeat_docs.py all
  python backend/scripts/parse_aeat_docs.py xsd
  python backend/scripts/parse_aeat_docs.py xls
  python backend/scripts/parse_aeat_docs.py verifactu
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any
import xml.etree.ElementTree as ET

# ---------------------------------------------------------------------------
# Project paths
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent  # TaxIA/

RENTA_XSD = PROJECT_ROOT / "docs/aeat/Renta-2025/Renta2024.xsd"
MODELO_130_XLS = PROJECT_ROOT / "docs/aeat/modelo-130-2026/DR130e15v12.xls"
MODELO_131_XLSX = PROJECT_ROOT / "docs/aeat/modelo-130-2026/DR131_2026.xlsx"
VERIFACTU_DIR = PROJECT_ROOT / "docs/aeat/VeriFactu"
REFERENCE_DIR = PROJECT_ROOT / "data/reference"

# XSD namespaces
XS_NS = "http://www.w3.org/2001/XMLSchema"
XS = f"{{{XS_NS}}}"

# ---------------------------------------------------------------------------
# Helper: strip namespace from a tag
# ---------------------------------------------------------------------------

def _local(tag: str) -> str:
    """Return local name without namespace."""
    return tag.split("}")[-1] if "}" in tag else tag


# ---------------------------------------------------------------------------
# A. XSD → JSON (Renta2024.xsd)
# ---------------------------------------------------------------------------

def parse_xsd(xsd_path: Path, output_path: Path) -> dict[str, Any]:
    """
    Parse Renta2024.xsd into a structured JSON reference file.

    Strategy:
    - Read raw text with ISO-8859-1 encoding to extract <!-- Página N --> comments
      and map line numbers to section names.
    - Parse the XML tree to collect:
        * simpleType definitions → type_catalog (name → restrictions dict)
        * complexType definitions → complex_types (name → list of child elements)
        * element declarations at the top level
    - Assign each complexType/element to a Pagina section based on comment proximity.
    - Emit grouped JSON.

    Args:
        xsd_path: Path to Renta2024.xsd
        output_path: Destination JSON path

    Returns:
        Summary dict for printing.
    """
    print(f"  Reading {xsd_path} ...")
    raw = xsd_path.read_bytes()
    # Decode with latin-1 (superset of ISO-8859-1); replace errors so we keep text
    text = raw.decode("latin-1", errors="replace")
    lines = text.splitlines()

    # --- Step 1: locate <!-- Página N --> comment markers by line number ---
    page_comment_re = re.compile(
        r"<!--\s*P[aá]gina\s+(\d+)", re.IGNORECASE
    )
    # Map: line_index → section_label (e.g. "Pagina01")
    section_markers: list[tuple[int, str]] = []
    for idx, line in enumerate(lines):
        m = page_comment_re.search(line)
        if m:
            num = int(m.group(1))
            label = f"Pagina{num:02d}"
            section_markers.append((idx, label))

    # Build a function: given a line number → nearest section label
    def section_for_line(lineno: int) -> str:
        """Return the section label for a given 0-based line index."""
        result = "Pagina00"  # before first marker
        for marker_line, label in section_markers:
            if marker_line <= lineno:
                result = label
            else:
                break
        return result

    # --- Step 2: parse the XML tree ---
    # ET doesn't handle encoding declarations well when parsing bytes directly
    # for ISO-8859-1; parse the decoded text as UTF-8 after re-encoding
    # We need to strip the XML declaration to re-parse cleanly.
    clean_text = re.sub(r"<\?xml[^>]+\?>", "", text, count=1)
    try:
        root = ET.fromstring(clean_text.encode("utf-8", errors="replace"))
    except ET.ParseError as exc:
        print(f"  WARNING: XML parse error: {exc}. Attempting lenient parse...")
        # Fall back: remove problematic chars from patterns
        clean_text = re.sub(r'value="[^"]*[^\x00-\x7F][^"]*"', 'value="..."', clean_text)
        root = ET.fromstring(clean_text.encode("utf-8", errors="replace"))

    # Detect namespace prefix used (xs: or default)
    # The tag of root will tell us the namespace
    ns_map: dict[str, str] = {}
    # Collect all namespaces from the raw text
    ns_decl_re = re.compile(r'xmlns(?::(\w+))?="([^"]+)"')
    for prefix, uri in ns_decl_re.findall(text[:2000]):
        ns_map[prefix or ""] = uri

    xs_prefix = ""
    for prefix, uri in ns_map.items():
        if uri == XS_NS:
            xs_prefix = prefix
            break

    # Determine the Clark notation prefix to use
    # If xs_prefix == "xs", elements are {XS_NS}simpleType etc.
    # If default ns is XS_NS, elements are {XS_NS}simpleType
    clark = XS  # always use Clark notation from ET

    # --- Step 3: build type_catalog from simpleType definitions ---
    type_catalog: dict[str, dict[str, Any]] = {}

    for st in root.findall(f"{clark}simpleType"):
        type_name = st.get("name", "")
        if not type_name:
            continue
        restrictions: dict[str, Any] = {}
        doc_text = ""

        # Extract documentation
        for doc in st.findall(f".//{clark}documentation"):
            if doc.text:
                doc_text = " ".join(doc.text.split())

        # Extract restriction details
        for restr in st.findall(f".//{clark}restriction"):
            base = restr.get("base", "")
            restrictions["base"] = base.split(":")[-1]  # strip prefix

            enums = [e.get("value", "") for e in restr.findall(f"{clark}enumeration")]
            if enums:
                restrictions["enumeration"] = enums

            for facet in ["maxLength", "minLength", "pattern", "minInclusive",
                          "maxInclusive", "fractionDigits", "totalDigits", "whiteSpace"]:
                el = restr.find(f"{clark}{facet}")
                if el is not None:
                    restrictions[facet] = el.get("value", "")

        type_catalog[type_name] = {
            "restrictions": restrictions,
            "documentation": doc_text[:300] if doc_text else "",
        }

    # --- Step 4: collect complexType definitions and their child elements ---
    # We need line numbers for each complexType to assign to a section.
    # ET doesn't expose line numbers reliably for all elements, so we
    # use a regex pass over the raw text to map type names → line numbers.
    type_lineno: dict[str, int] = {}
    ct_name_re = re.compile(r'<xs:complexType\s+name="([^"]+)"')
    for idx, line in enumerate(lines):
        m = ct_name_re.search(line)
        if m:
            type_lineno[m.group(1)] = idx

    # Also for simpleType
    st_name_re = re.compile(r'<xs:simpleType\s+name="([^"]+)"')
    for idx, line in enumerate(lines):
        m = st_name_re.search(line)
        if m:
            type_lineno[m.group(1)] = idx

    def _resolve_restrictions(type_ref: str) -> dict[str, Any]:
        """Look up a type name in type_catalog and return its restrictions."""
        # Strip namespace prefix if present
        bare = type_ref.split(":")[-1]
        entry = type_catalog.get(bare, {})
        return entry.get("restrictions", {})

    def _collect_fields(ct_elem: ET.Element, path_prefix: str, depth: int = 0) -> list[dict[str, Any]]:
        """
        Recursively collect element declarations from a complexType element.

        Args:
            ct_elem: The complexType ET element to walk
            path_prefix: XPath prefix for building field paths
            depth: Recursion guard (max 5 levels)

        Returns:
            List of field dicts.
        """
        if depth > 5:
            return []
        fields: list[dict[str, Any]] = []

        for child in ct_elem.iter(f"{clark}element"):
            elem_name = child.get("name")
            if not elem_name:
                continue
            type_ref = child.get("type", "")
            bare_type = type_ref.split(":")[-1] if type_ref else "complexType"
            field_path = f"{path_prefix}/{elem_name}"

            restrictions = _resolve_restrictions(type_ref) if type_ref else {}

            field: dict[str, Any] = {
                "name": elem_name,
                "path": field_path,
                "type": bare_type,
                "min_occurs": child.get("minOccurs", "1"),
                "max_occurs": child.get("maxOccurs", "1"),
            }
            if restrictions:
                # Summarize key restriction info
                if "maxLength" in restrictions:
                    field["max_length"] = int(restrictions["maxLength"])
                if "minInclusive" in restrictions:
                    field["min_value"] = restrictions["minInclusive"]
                if "maxInclusive" in restrictions:
                    field["max_value"] = restrictions["maxInclusive"]
                if "enumeration" in restrictions:
                    vals = restrictions["enumeration"]
                    field["allowed_values"] = vals if len(vals) <= 20 else vals[:20] + ["..."]
                if "pattern" in restrictions:
                    field["pattern"] = restrictions["pattern"][:100]
                if "base" in restrictions:
                    field["base_type"] = restrictions["base"]

            # Include documentation if present on the element itself
            doc_el = child.find(f".//{clark}documentation")
            if doc_el is not None and doc_el.text:
                field["description"] = " ".join(doc_el.text.split())[:200]

            fields.append(field)

        return fields

    # --- Step 5: group complexTypes into sections ---
    sections: dict[str, dict[str, Any]] = {}

    # Section description map (based on known Renta structure)
    section_descriptions = {
        "Pagina00": "Tipos basicos y definiciones comunes",
        "Pagina01": "Datos identificativos del declarante",
        "Pagina02": "Datos personales y familiares",
        "Pagina03": "Autoliquidacion rectificativa",
        "Pagina04": "Rendimientos del trabajo",
        "Pagina05": "Rendimientos del capital inmobiliario",
        "Pagina06": "Rendimientos del capital mobiliario",
        "Pagina07": "Rendimientos de actividades economicas",
        "Pagina08": "Ganancias y perdidas patrimoniales",
        "Pagina09": "Base imponible del ahorro",
        "Pagina10": "Base liquidable y minimo personal",
        "Pagina11": "Cuota integra",
        "Pagina12": "Deducciones de la cuota",
        "Pagina13": "Cuota liquida y resultado",
        "Pagina14": "Informacion adicional",
        "Pagina15": "Datos complementarios",
        "Pagina16": "Anexos e informacion adicional avanzada",
    }

    all_element_count = 0

    for ct in root.findall(f"{clark}complexType"):
        ct_name = ct.get("name", "")
        if not ct_name:
            continue
        lineno = type_lineno.get(ct_name, 0)
        section = section_for_line(lineno)

        if section not in sections:
            sections[section] = {
                "description": section_descriptions.get(section, section),
                "types": [],
                "fields": [],
            }

        # Collect direct child elements
        base_path = f"/Declaracion/{section}"
        fields = _collect_fields(ct, base_path, depth=0)
        all_element_count += len(fields)

        # Store the type definition
        sections[section]["types"].append({
            "name": ct_name,
            "field_count": len(fields),
        })
        sections[section]["fields"].extend(fields)

    # Also handle top-level element declarations
    for el in root.findall(f"{clark}element"):
        el_name = el.get("name", "")
        if not el_name:
            continue
        # Try to find its line
        lineno = 0
        for idx, line in enumerate(lines):
            if f'name="{el_name}"' in line and "xs:element" in line:
                lineno = idx
                break
        section = section_for_line(lineno)
        if section not in sections:
            sections[section] = {
                "description": section_descriptions.get(section, section),
                "types": [],
                "fields": [],
            }
        # Add as a top-level field
        type_ref = el.get("type", "complexType")
        sections[section]["fields"].append({
            "name": el_name,
            "path": f"/Declaracion/{el_name}",
            "type": type_ref.split(":")[-1],
            "is_root_element": True,
        })
        all_element_count += 1

    # Sort sections
    sorted_sections = dict(sorted(sections.items()))

    output = {
        "year": 2024,
        "source_file": "Renta2024.xsd",
        "encoding": "ISO-8859-1",
        "total_complex_types": len(type_catalog),
        "total_elements": all_element_count,
        "simple_type_catalog": {
            name: {
                "base": info["restrictions"].get("base", "string"),
                "max_length": info["restrictions"].get("maxLength"),
                "pattern": info["restrictions"].get("pattern", "")[:80] if info["restrictions"].get("pattern") else None,
                "enumeration": info["restrictions"].get("enumeration", [])[:20] if "enumeration" in info["restrictions"] else None,
                "min": info["restrictions"].get("minInclusive"),
                "max": info["restrictions"].get("maxInclusive"),
                "description": info["documentation"][:150] if info["documentation"] else None,
            }
            for name, info in type_catalog.items()
        },
        "sections": sorted_sections,
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(output, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    return {
        "output": str(output_path),
        "sections": len(sorted_sections),
        "simple_types": len(type_catalog),
        "elements": all_element_count,
    }


# ---------------------------------------------------------------------------
# B. XLS/XLSX → JSON (Modelo 130 / 131)
# ---------------------------------------------------------------------------

def _parse_xls_sheet(ws: Any, sheet_name: str) -> dict[str, Any]:
    """
    Parse a single worksheet from a Modelo XLS/XLSX file.

    Expects columns: Nº, Posic., Lon, Tipo, Descripción, Validación, Contenido
    Tries to auto-detect header row.

    Args:
        ws: openpyxl worksheet or xlrd sheet object
        sheet_name: Name of the sheet for logging

    Returns:
        Dict with 'description' and 'fields' list.
    """
    # Detect whether this is an xlrd sheet or openpyxl worksheet
    is_xlrd = hasattr(ws, "row_values")

    def get_cell(row_idx: int, col_idx: int) -> Any:
        if is_xlrd:
            try:
                return ws.cell_value(row_idx, col_idx)
            except IndexError:
                return None
        else:
            cell = ws.cell(row=row_idx + 1, column=col_idx + 1)
            return cell.value

    def num_rows() -> int:
        if is_xlrd:
            return ws.nrows
        return ws.max_row or 0

    # Find header row — look for a row with "Nº" or "No" or "NUM" in col 0 or 1
    header_row = 0
    for r in range(min(20, num_rows())):
        for c in range(min(5, 10)):
            val = get_cell(r, c)
            if val is None:
                continue
            s = str(val).strip().upper()
            if s in ("Nº", "NO", "NUM", "N", "#"):
                header_row = r
                break
        else:
            continue
        break

    # Find column mapping from header row
    col_map: dict[str, int] = {}
    keywords = {
        "NUM": ["nº", "no", "num", "#", "n"],
        "POSIC": ["posic", "pos", "posición", "posicion"],
        "LON": ["lon", "long", "longitud", "length"],
        "TIPO": ["tipo", "type", "tip"],
        "DESC": ["descripción", "descripcion", "desc", "nombre", "campo"],
        "VALID": ["validación", "validacion", "valid", "reglas"],
        "CONTENIDO": ["contenido", "valor", "content"],
    }
    max_cols = 20
    for c in range(max_cols):
        val = get_cell(header_row, c)
        if val is None:
            continue
        s = str(val).strip().lower()
        for key, patterns in keywords.items():
            if key not in col_map and any(p in s for p in patterns):
                col_map[key] = c
                break

    fields: list[dict[str, Any]] = []
    for r in range(header_row + 1, num_rows()):
        num_val = get_cell(r, col_map.get("NUM", 0))
        if num_val is None or str(num_val).strip() == "":
            continue
        # Skip rows where num is text header (re-occurrence of header)
        try:
            field_num = int(float(str(num_val)))
        except (ValueError, TypeError):
            continue

        def safe_get(key: str, col_default: int = 99) -> Any:
            col = col_map.get(key, col_default)
            if col >= max_cols:
                return None
            v = get_cell(r, col)
            if v is None:
                return None
            if isinstance(v, float) and v == int(v):
                return int(v)
            return str(v).strip() if isinstance(v, str) else v

        field: dict[str, Any] = {
            "num": field_num,
            "position": safe_get("POSIC"),
            "length": safe_get("LON"),
            "type": safe_get("TIPO"),
            "description": safe_get("DESC"),
            "validation": safe_get("VALID"),
            "content": safe_get("CONTENIDO"),
        }
        # Remove None values to keep output clean
        field = {k: v for k, v in field.items() if v is not None}
        fields.append(field)

    return {
        "description": sheet_name,
        "header_row_detected": header_row,
        "fields": fields,
    }


def parse_xls(
    xls_path: Path,
    xlsx_path: Path,
    out_130: Path,
    out_131: Path,
) -> dict[str, Any]:
    """
    Parse Modelo 130 (XLS) and 131 (XLSX) into JSON reference files.

    Args:
        xls_path: Path to DR130e15v12.xls
        xlsx_path: Path to DR131_2026.xlsx
        out_130: Output path for modelo_130_fields.json
        out_131: Output path for modelo_131_fields.json

    Returns:
        Summary dict.
    """
    results: dict[str, Any] = {}

    # -- Modelo 130 (XLS via xlrd) --
    print(f"  Reading {xls_path} (xlrd) ...")
    try:
        import xlrd  # type: ignore
        wb = xlrd.open_workbook(str(xls_path))
        pages_130: dict[str, Any] = {}
        version = "unknown"

        for sheet_idx in range(wb.nsheets):
            ws = wb.sheet_by_index(sheet_idx)
            sheet_name = wb.sheet_names()[sheet_idx]
            # Try to detect version from sheet name or cell
            if sheet_idx == 0:
                for r in range(min(10, ws.nrows)):
                    for c in range(min(8, ws.ncols)):
                        cell_val = str(ws.cell_value(r, c))
                        vmatch = re.search(r"v(\d+\.\d+)", cell_val, re.IGNORECASE)
                        if vmatch:
                            version = vmatch.group(1)
                            break

            page_key = f"page_{sheet_idx}"
            page_data = _parse_xls_sheet(ws, sheet_name)
            pages_130[page_key] = page_data

        output_130 = {
            "modelo": "130",
            "description": "Estimacion Directa — Pago Fraccionado IRPF",
            "version": version,
            "source_file": xls_path.name,
            "total_sheets": wb.nsheets,
            "pages": pages_130,
        }
        out_130.parent.mkdir(parents=True, exist_ok=True)
        out_130.write_text(
            json.dumps(output_130, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        total_fields_130 = sum(len(p["fields"]) for p in pages_130.values())
        results["130"] = {
            "output": str(out_130),
            "sheets": wb.nsheets,
            "fields": total_fields_130,
        }
        print(f"    -> {wb.nsheets} sheets, {total_fields_130} fields")

    except ImportError:
        print("  WARNING: xlrd not installed. Skipping Modelo 130.")
        results["130"] = {"error": "xlrd not installed"}
    except Exception as exc:
        print(f"  ERROR parsing Modelo 130: {exc}")
        results["130"] = {"error": str(exc)}

    # -- Modelo 131 (XLSX via openpyxl) --
    print(f"  Reading {xlsx_path} (openpyxl) ...")
    try:
        import openpyxl  # type: ignore
        wb2 = openpyxl.load_workbook(str(xlsx_path), data_only=True)
        pages_131: dict[str, Any] = {}
        version2 = "unknown"

        for sheet_idx, ws in enumerate(wb2.worksheets):
            sheet_name = ws.title or f"Sheet{sheet_idx}"
            # Try to detect version from the first few cells
            if sheet_idx == 0:
                for r in range(1, min(10, (ws.max_row or 1) + 1)):
                    for c in range(1, min(9, (ws.max_column or 1) + 1)):
                        cell_val = ws.cell(row=r, column=c).value
                        if cell_val and isinstance(cell_val, str):
                            vmatch = re.search(r"v(\d+\.\d+)", cell_val, re.IGNORECASE)
                            if vmatch:
                                version2 = vmatch.group(1)

            page_key = f"page_{sheet_idx}"
            page_data = _parse_xls_sheet(ws, sheet_name)
            pages_131[page_key] = page_data

        output_131 = {
            "modelo": "131",
            "description": "Estimacion Objetiva (Modulos) — Pago Fraccionado IRPF",
            "version": version2,
            "source_file": xlsx_path.name,
            "total_sheets": len(wb2.worksheets),
            "pages": pages_131,
        }
        out_131.parent.mkdir(parents=True, exist_ok=True)
        out_131.write_text(
            json.dumps(output_131, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        total_fields_131 = sum(len(p["fields"]) for p in pages_131.values())
        results["131"] = {
            "output": str(out_131),
            "sheets": len(wb2.worksheets),
            "fields": total_fields_131,
        }
        print(f"    -> {len(wb2.worksheets)} sheets, {total_fields_131} fields")

    except ImportError:
        print("  WARNING: openpyxl not installed. Skipping Modelo 131.")
        results["131"] = {"error": "openpyxl not installed"}
    except Exception as exc:
        print(f"  ERROR parsing Modelo 131: {exc}")
        results["131"] = {"error": str(exc)}

    return results


# ---------------------------------------------------------------------------
# C. VeriFactu XSD/WSDL → human-readable .txt
# ---------------------------------------------------------------------------

def _extract_xsd_as_text(xsd_path: Path, summary: str) -> str:
    """
    Convert an XSD file into a human-readable text document for RAG indexing.

    Extracts:
    - All complexType definitions with their child elements and documentation
    - All simpleType definitions with their restrictions
    - Top-level element declarations

    Args:
        xsd_path: Path to the XSD file
        summary: One-line summary to include at the top of the document

    Returns:
        Formatted text content.
    """
    text = xsd_path.read_text(encoding="utf-8", errors="replace")
    lines_raw = text.splitlines()

    # Clean XML declaration for parsing
    clean = re.sub(r"<\?xml[^>]+\?>", "", text, count=1)
    try:
        root = ET.fromstring(clean)
    except ET.ParseError:
        clean = re.sub(r'value="[^"]*"', 'value="..."', clean)
        try:
            root = ET.fromstring(clean)
        except ET.ParseError as exc2:
            return f"[ERROR parsing {xsd_path.name}: {exc2}]\n"

    # Detect Clark notation for this file
    # Some use default namespace (no prefix), some use xs:
    root_tag = root.tag
    if root_tag.startswith("{"):
        ns = root_tag.split("}")[0][1:]
        clark = f"{{{ns}}}"
    else:
        clark = ""
        ns = ""

    def ftag(local: str) -> str:
        return f"{clark}{local}"

    output_parts: list[str] = []
    output_parts.append("=" * 70)
    output_parts.append(f"FICHERO: {xsd_path.name}")
    output_parts.append(f"NAMESPACE: {ns or '(default)'}")
    output_parts.append(summary)
    output_parts.append("=" * 70)
    output_parts.append("")

    # --- Top-level elements ---
    top_elements = root.findall(ftag("element"))
    if top_elements:
        output_parts.append("ELEMENTOS RAIZ (Top-level elements)")
        output_parts.append("-" * 40)
        for el in top_elements:
            name = el.get("name", "?")
            etype = el.get("type", "(inline)")
            doc_el = el.find(f".//{ftag('documentation')}")
            doc = ""
            if doc_el is not None and doc_el.text:
                doc = " ".join(doc_el.text.split())[:300]
            output_parts.append(f"  Elemento: {name}")
            output_parts.append(f"  Tipo    : {etype}")
            if doc:
                output_parts.append(f"  Desc    : {doc}")
            output_parts.append("")

    # --- complexType definitions ---
    complex_types = root.findall(ftag("complexType"))
    if complex_types:
        output_parts.append("")
        output_parts.append("TIPOS COMPLEJOS (complexType)")
        output_parts.append("=" * 50)

        for ct in complex_types:
            ct_name = ct.get("name", "(anonimo)")
            doc_el = ct.find(f".//{ftag('documentation')}")
            ct_doc = ""
            if doc_el is not None and doc_el.text:
                ct_doc = " ".join(doc_el.text.split())[:400]

            output_parts.append(f"\nTIPO: {ct_name}")
            if ct_doc:
                output_parts.append(f"  Descripcion: {ct_doc}")

            # Collect child elements
            child_elems = ct.findall(f".//{ftag('element')}")
            if child_elems:
                output_parts.append("  Campos:")
                for child in child_elems:
                    child_name = child.get("name", "?")
                    child_type = child.get("type", "(inline)")
                    min_o = child.get("minOccurs", "1")
                    max_o = child.get("maxOccurs", "1")
                    child_doc_el = child.find(f".//{ftag('documentation')}")
                    child_doc = ""
                    if child_doc_el is not None and child_doc_el.text:
                        child_doc = " ".join(child_doc_el.text.split())[:200]
                    required = "obligatorio" if min_o != "0" else "opcional"
                    output_parts.append(
                        f"    - {child_name}: {child_type} [{required}, maxOccurs={max_o}]"
                    )
                    if child_doc:
                        output_parts.append(f"      Desc: {child_doc}")

            # Check for extensions
            ext = ct.find(f".//{ftag('extension')}")
            if ext is not None:
                output_parts.append(f"  Extiende: {ext.get('base', '?')}")

            output_parts.append("")

    # --- simpleType definitions ---
    simple_types = root.findall(ftag("simpleType"))
    if simple_types:
        output_parts.append("")
        output_parts.append("TIPOS SIMPLES (simpleType)")
        output_parts.append("=" * 50)

        for st in simple_types:
            st_name = st.get("name", "(anonimo)")
            doc_el = st.find(f".//{ftag('documentation')}")
            st_doc = ""
            if doc_el is not None and doc_el.text:
                st_doc = " ".join(doc_el.text.split())[:300]

            restr = st.find(f".//{ftag('restriction')}")
            if restr is None:
                continue

            base = restr.get("base", "string")
            enums = [e.get("value", "") for e in restr.findall(ftag("enumeration"))]
            max_len = restr.find(ftag("maxLength"))
            pattern = restr.find(ftag("pattern"))

            output_parts.append(f"\nTIPO SIMPLE: {st_name} (base: {base})")
            if st_doc:
                output_parts.append(f"  Descripcion: {st_doc}")
            if max_len is not None:
                output_parts.append(f"  Longitud maxima: {max_len.get('value')}")
            if pattern is not None:
                output_parts.append(f"  Patron: {pattern.get('value', '')[:100]}")
            if enums:
                output_parts.append(f"  Valores permitidos: {', '.join(enums[:30])}")
                if len(enums) > 30:
                    output_parts.append(f"  ... y {len(enums) - 30} valores mas")

    return "\n".join(output_parts) + "\n"


def _extract_wsdl_as_text(wsdl_path: Path) -> str:
    """
    Convert a WSDL file into a human-readable text document for RAG indexing.

    Extracts:
    - Service name, target namespace
    - Port types and operations (with input/output message names)
    - Binding info (style, transport)
    - Service ports with endpoint URLs

    Args:
        wsdl_path: Path to the WSDL file

    Returns:
        Formatted text content.
    """
    text = wsdl_path.read_text(encoding="utf-8", errors="replace")
    clean = re.sub(r"<\?xml[^>]+\?>", "", text, count=1)

    try:
        root = ET.fromstring(clean)
    except ET.ParseError as exc:
        return f"[ERROR parsing {wsdl_path.name}: {exc}]\n"

    WSDL_NS = "http://schemas.xmlsoap.org/wsdl/"
    SOAP_NS = "http://schemas.xmlsoap.org/wsdl/soap/"
    wsdl = f"{{{WSDL_NS}}}"
    soap = f"{{{SOAP_NS}}}"

    tns = root.get("targetNamespace", "")
    service_name = root.get("name", wsdl_path.stem)

    output_parts: list[str] = []
    output_parts.append("=" * 70)
    output_parts.append(f"FICHERO: {wsdl_path.name}")
    output_parts.append(f"SERVICIO SOAP: {service_name}")
    output_parts.append(f"TARGET NAMESPACE: {tns}")
    output_parts.append("=" * 70)
    output_parts.append("")
    output_parts.append(
        "Este fichero WSDL define el servicio web SOAP del sistema VeriFactu de la AEAT "
        "para la gestion de registros de facturacion verificables (Ley Antifraude 2021)."
    )
    output_parts.append("")

    # Messages
    messages = root.findall(f"{wsdl}message")
    if messages:
        output_parts.append("MENSAJES SOAP")
        output_parts.append("-" * 40)
        for msg in messages:
            msg_name = msg.get("name", "?")
            parts = msg.findall(f"{wsdl}part")
            for part in parts:
                part_name = part.get("name", "?")
                elem = part.get("element", "?")
                output_parts.append(f"  Mensaje: {msg_name}")
                output_parts.append(f"    Part : {part_name} → elemento: {elem}")
        output_parts.append("")

    # Port types
    port_types = root.findall(f"{wsdl}portType")
    if port_types:
        output_parts.append("OPERACIONES DISPONIBLES (portType)")
        output_parts.append("-" * 40)
        for pt in port_types:
            pt_name = pt.get("name", "?")
            output_parts.append(f"\n  PortType: {pt_name}")
            for op in pt.findall(f"{wsdl}operation"):
                op_name = op.get("name", "?")
                inp = op.find(f"{wsdl}input")
                out = op.find(f"{wsdl}output")
                inp_msg = inp.get("message", "?") if inp is not None else "?"
                out_msg = out.get("message", "?") if out is not None else "?"
                output_parts.append(f"    Operacion: {op_name}")
                output_parts.append(f"      Input  : {inp_msg}")
                output_parts.append(f"      Output : {out_msg}")
        output_parts.append("")

    # Bindings
    bindings = root.findall(f"{wsdl}binding")
    if bindings:
        output_parts.append("BINDINGS SOAP")
        output_parts.append("-" * 40)
        for binding in bindings:
            b_name = binding.get("name", "?")
            b_type = binding.get("type", "?")
            soap_b = binding.find(f"{soap}binding")
            style = soap_b.get("style", "?") if soap_b is not None else "?"
            transport = soap_b.get("transport", "?") if soap_b is not None else "?"
            output_parts.append(f"\n  Binding: {b_name} → {b_type}")
            output_parts.append(f"    Estilo    : {style}")
            output_parts.append(f"    Transporte: {transport}")
            for op in binding.findall(f"{wsdl}operation"):
                op_name = op.get("name", "?")
                soap_op = op.find(f"{soap}operation")
                action = soap_op.get("soapAction", "") if soap_op is not None else ""
                output_parts.append(f"    Operacion: {op_name} (soapAction='{action}')")
        output_parts.append("")

    # Services and endpoints
    services = root.findall(f"{wsdl}service")
    if services:
        output_parts.append("ENDPOINTS (URLs de acceso)")
        output_parts.append("-" * 40)
        for svc in services:
            svc_name = svc.get("name", "?")
            output_parts.append(f"\n  Servicio: {svc_name}")
            for port in svc.findall(f"{wsdl}port"):
                port_name = port.get("name", "?")
                binding = port.get("binding", "?")
                soap_addr = port.find(f"{{{SOAP_NS}}}address")
                location = soap_addr.get("location", "?") if soap_addr is not None else "?"
                output_parts.append(f"    Puerto  : {port_name}")
                output_parts.append(f"    Binding : {binding}")
                output_parts.append(f"    URL     : {location}")
        output_parts.append("")

    return "\n".join(output_parts) + "\n"


def parse_verifactu(verifactu_dir: Path) -> dict[str, Any]:
    """
    Convert all VeriFactu XSD and WSDL files into human-readable .txt companions.

    Args:
        verifactu_dir: Directory containing VeriFactu schema files

    Returns:
        Summary dict mapping filename → output path.
    """
    summaries = {
        "SuministroInformacion.xsd": (
            "Define los tipos comunes del sistema VeriFactu: cabecera de envio, "
            "datos de presentacion, identificadores de factura, tipos de factura, "
            "datos del emisor, y tipos de firma electronica para el registro de "
            "facturacion verificable (LIRPF Art. 29)."
        ),
        "SuministroLR.xsd": (
            "Define la estructura del mensaje de envio de registros de facturacion "
            "al sistema VeriFactu de la AEAT. Contiene el elemento raiz "
            "RegFactuSistemaFacturacion con cabecera y lista de RegistroFactura "
            "(alta o anulacion)."
        ),
        "RespuestaSuministro.xsd": (
            "Define la estructura de la respuesta AEAT a un envio de registros de "
            "facturacion VeriFactu. Incluye CSV del envio, estado global "
            "(Correcto/Incorrecto/Parcialmente correcto) y estado por linea."
        ),
        "ConsultaLR.xsd": (
            "Define la estructura para consultar registros de facturacion en el "
            "sistema VeriFactu de la AEAT. Permite filtrar por periodo, numero de "
            "factura, contraparte y otros criterios."
        ),
        "SistemaFacturacion.wsdl": None,  # Special WSDL handling
    }

    results: dict[str, Any] = {}

    for filename, summary in summaries.items():
        src = verifactu_dir / filename
        if not src.exists():
            print(f"  WARNING: {src} not found, skipping.")
            continue

        stem = src.stem
        out_path = verifactu_dir / f"{stem}_reference.txt"

        print(f"  Processing {filename} ...")

        if filename.endswith(".wsdl"):
            content = _extract_wsdl_as_text(src)
        else:
            content = _extract_xsd_as_text(src, summary or "")

        out_path.write_text(content, encoding="utf-8")
        line_count = content.count("\n")
        print(f"    -> {out_path.name} ({line_count} lines)")

        results[filename] = {
            "output": str(out_path),
            "lines": line_count,
        }

    return results


# ---------------------------------------------------------------------------
# Main CLI
# ---------------------------------------------------------------------------

def main() -> None:
    """Entry point with argparse subcommands."""
    parser = argparse.ArgumentParser(
        description="Parse AEAT documents into structured reference files for TaxIA."
    )
    parser.add_argument(
        "command",
        choices=["xsd", "xls", "verifactu", "all"],
        help="Which parser to run",
    )
    args = parser.parse_args()

    print("\nTaxIA — AEAT Document Parser")
    print("=" * 50)

    if args.command in ("xsd", "all"):
        print("\n[A] Parsing Renta2024.xsd ...")
        result = parse_xsd(
            xsd_path=RENTA_XSD,
            output_path=REFERENCE_DIR / "renta_2024_schema.json",
        )
        print(
            f"  Done: {result['sections']} sections, "
            f"{result['simple_types']} simple types, "
            f"{result['elements']} elements"
        )
        print(f"  Output: {result['output']}")

    if args.command in ("xls", "all"):
        print("\n[B] Parsing Modelo 130 / 131 XLS/XLSX ...")
        result = parse_xls(
            xls_path=MODELO_130_XLS,
            xlsx_path=MODELO_131_XLSX,
            out_130=REFERENCE_DIR / "modelo_130_fields.json",
            out_131=REFERENCE_DIR / "modelo_131_fields.json",
        )
        for modelo, info in result.items():
            if "error" in info:
                print(f"  Modelo {modelo}: ERROR — {info['error']}")
            else:
                print(f"  Modelo {modelo}: {info['fields']} fields → {info['output']}")

    if args.command in ("verifactu", "all"):
        print("\n[C] Parsing VeriFactu XSD/WSDL ...")
        result = parse_verifactu(VERIFACTU_DIR)
        for fname, info in result.items():
            if "error" in info:
                print(f"  {fname}: ERROR — {info['error']}")
            else:
                print(f"  {fname}: {info['lines']} lines → {info['output']}")

    print("\nDone.")


if __name__ == "__main__":
    main()
